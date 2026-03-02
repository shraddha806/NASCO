import os
import re
import sys
import time
import asyncio
import json
import pandas as pd
from datetime import datetime
from imap_tools import MailBox, AND
from dotenv import load_dotenv
from browser_use import (
    Agent,
    ChatBrowserUse,
    Tools,
    ActionResult,
    BrowserSession,
)

import warnings
warnings.filterwarnings("ignore", category=ResourceWarning)

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

load_dotenv()

BASE_URL = os.getenv("BASE_URL", "https://www.icicilombard.com/")
VALID_USERNAME = os.getenv("VALID_USERNAME", "6362493807")
VEHICLE_REG_NO = os.getenv("Vehicle_Registration_Number", "TS07KE1255")
EMAIL_ID = os.getenv("EMAIL_ID", "priyankashah8324@gmail.com")
FULL_NAME = os.getenv("FULL_NAME", "Priyanka Shah")
BROWSER_USE_API_KEY = os.getenv("BROWSER_USE_API_KEY")

GMAIL_EMAIL = os.getenv("EMAIL_ADDRESS")
GMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
IMAP_HOST = "imap.gmail.com"

OTP_TIMEOUT = 120
OTP_INTERVAL = 5

def extract_otp(text: str):
    patterns = [
        r"verification\s+OTP\s+is\s+(\d{6})",
        r"OTP\s+(\d{6})",
        r"Your\s+(\d{6})",
        r"\b\d{6}\b"
    ]
    for pat in patterns:
        match = re.search(pat, text, re.IGNORECASE)
        if match:
            otp = match.group(1)
            print(f"DEBUG: Matched '{pat}' -> OTP: {otp}")
            return otp
    return None

def blocking_fetch_otp() -> str:
    print("DEBUG: Searching for new OTP email...")
    start = time.time()

    while time.time() - start < OTP_TIMEOUT:
        try:
            with MailBox(IMAP_HOST).login(GMAIL_EMAIL, GMAIL_PASSWORD, "INBOX") as mailbox:
                for msg in mailbox.fetch(AND(seen=False), reverse=True, limit=5):
                    body = msg.text or msg.html or ""
                    subject = msg.subject.lower()
                    
                    if "icici" in subject or "icici" in body.lower():
                        otp = extract_otp(body)
                        if otp and len(otp) == 6:
                            mailbox.flag([msg.uid], ["\\Seen"], True)
                            print(f"DEBUG: Successfully fetched OTP: {otp}")
                            return otp
        except Exception as e:
            print(f"DEBUG: IMAP error: {e}")
        time.sleep(OTP_INTERVAL)
    
    raise RuntimeError("OTP not received within timeout period")

tools = Tools()

@tools.action(description="Fetch the 6-digit ICICI Lombard OTP from Gmail inbox")
async def fetch_otp_from_email(browser_session: BrowserSession) -> ActionResult:
    """Fetches latest UNSEEN ICICI OTP email and returns exact 6-digit code."""
    loop = asyncio.get_running_loop()
    try:
        otp = await loop.run_in_executor(None, blocking_fetch_otp)
        return ActionResult(
            extracted_content=f"OTP '{otp}' fetched from ICICI email. Enter digits 1-6 into OTP fields.",
            long_term_memory=f"Current OTP for {VALID_USERNAME} is {otp}"
        )
    except Exception as e:
        return ActionResult(
            error=f"Failed to fetch OTP: {str(e)}",
            long_term_memory="OTP fetch failed - check Gmail credentials"
        )

async def run_icici_lombard():
    """Run ICICI Lombard insurance automation"""
    print("\n" + "=" * 80)
    print("STARTING ICICI LOMBARD")
    print("=" * 80)
    
    agent = Agent(
        task=f"""
You are an insurance automation agent.

Steps:
1. Login to {BASE_URL} using mobile {VALID_USERNAME} and OTP.
2. Navigate to Car Insurance page.
3. Enter:
   - Vehicle Number: {VEHICLE_REG_NO}
   - Mobile: {6362493807}
   - Email: {EMAIL_ID} → Send OTP
4. Wait for 6 OTP input boxes to appear
5. Call `fetch_otp_from_email` tool to get REAL OTP from Gmail
6. After OTP fields appear, identify ALL visible OTP input boxes (exactly 6)
7. Sort OTP inputs from LEFT to RIGHT
8. For each digit in OTP:
   - Click the corresponding OTP input
   - Clear it
   - Type EXACTLY one digit
   - Wait 300ms before next digit
9. ONLY AFTER all 6 digits are entered, click Verify
10. Click 'Verify' button
11. Confirm login success
12. Ensure login is successful and homepage is fully loaded.
13. Directly navigate to the Car Insurance page by opening:
    https://www.icicilombard.com/car-insurance
14. Wait for the Car Insurance page to fully load (URL change or car insurance form visible).
15. Enter vehicle {VEHICLE_REG_NO}, mobile {VALID_USERNAME}, email {EMAIL_ID}
16. Get quotes
17. Extract first 2-3 visible plan cards
18. For each plan: plan name, premium (integer), key benefits

Return ONLY this JSON (no markdown, no extra text):
{{
  "status": "success",
  "vehicle": "{VEHICLE_REG_NO}",
  "quotes": [
    {{"plan_name": "Plan 1", "annual_premium_inr": 4465, "key_benefits": ["Benefit 1", "Benefit 2"]}},
    {{"plan_name": "Plan 2", "annual_premium_inr": 5563, "key_benefits": ["Benefit 1", "Benefit 2"]}}
  ],
  "summary": "Extracted X plans"
}}
""",
        llm=ChatBrowserUse(
            model="bu-latest"
        ),
        tools=tools
    )

    result = await agent.run()
    result_text = result.final_result()
    
    print("\n" + "=" * 80)
    print("ICICI LOMBARD RESULT:")
    print(result_text)
    print("=" * 80)
    return result_text

async def run_policybazaar():
    """Run PolicyBazaar insurance automation"""
    print("\n" + "=" * 80)
    print("STARTING POLICYBAZAAR")
    print("=" * 80)
    
    policybazaar_url = "https://policybazaar.com"
    
    agent = Agent(
        task=f"""
Extract PolicyBazaar car insurance quotes for {VEHICLE_REG_NO}.

1. Go to {policybazaar_url}
2. Enter vehicle number {VEHICLE_REG_NO}, then name {FULL_NAME} and mobile {VALID_USERNAME}
3. Get to quotes page
4. Extract ONLY first 3 visible policy cards
5. For each: provider, plan name, premium (integer), 3 benefits

Return ONLY this JSON (no extra text, no markdown):
{{
  "status": "success",
  "vehicle": "{VEHICLE_REG_NO}",
  "quotes": [
    {{"insurance_provider": "Provider1", "plan_name": "Plan1", "annual_premium_inr": 12500, "key_benefits": ["Benefit1", "Benefit2", "Benefit3"]}},
    {{"insurance_provider": "Provider2", "plan_name": "Plan2", "annual_premium_inr": 13000, "key_benefits": ["Benefit1", "Benefit2", "Benefit3"]}},
    {{"insurance_provider": "Provider3", "plan_name": "Plan3", "annual_premium_inr": 11800, "key_benefits": ["Benefit1", "Benefit2", "Benefit3"]}}
  ],
  "summary": "Extracted 3 policies"
}}
""",
        llm=ChatBrowserUse(
            model="bu-latest"
        )
    )

    result = await agent.run()
    result_text = result.final_result()
    
    print("\n" + "=" * 80)
    print("POLICYBAZAAR RESULT:")
    print(result_text)
    print("=" * 80)
    return result_text

async def run_insurancedekho():
    """Run InsuranceDekho insurance automation"""
    print("\n" + "=" * 80)
    print("STARTING INSURANCEDEKHO")
    print("=" * 80)
    
    agent = Agent(
        task=f"""
Extract InsuranceDekho car insurance quotes for {VEHICLE_REG_NO}.

STEPS:
1. Go to https://www.insurancedekho.com/car-insurance
2. Enter vehicle {VEHICLE_REG_NO}, mobile {VALID_USERNAME}
3. Next page: enter name {FULL_NAME}, click blue checkbox, submit
4. Handle any popups (select "No" for claims/ownership questions if asked)
5. Click "Comprehensive" filter if available
6. Wait for quotes to load fully
7. Extract 2-3 visible plan cards with: provider name, premium (integer), cashless garages count, IDV value, 2-3 key benefits

IMPORTANT OUTPUT FORMAT:
When you call the 'done' action, the 'text' parameter MUST contain ONLY the JSON below.
Do NOT write a summary like "Extracted car insurance quotes". 
Do NOT add any explanation before or after the JSON.
The 'text' field must start with {{ and end with }}.

EXACT FORMAT FOR done action text:
{{
  "status": "success",
  "vehicle": "{VEHICLE_REG_NO}",
  "quotes": [
    {{
      "insurance_provider": "Zuno",
      "plan_name": "Zuno Comprehensive",
      "annual_premium_inr": 3599,
      "idv": "141515",
      "cashless_garages": "90",
      "key_benefits": ["FREE vehicle pick up", "6 months repair warranty"]
    }},
    {{
      "insurance_provider": "Liberty",
      "plan_name": "Liberty Comprehensive",
      "annual_premium_inr": 3611,
      "idv": "113208",
      "cashless_garages": "3860",
      "key_benefits": ["Towing assistance", "Unlimited Claims"]
    }}
  ],
  "summary": "Extracted 2 plans from InsuranceDekho"
}}
""",
        llm=ChatBrowserUse(
            model="bu-latest"
        )
    )

    result = await agent.run()
    result_text = result.final_result()
    
    print("\n" + "=" * 80)
    print("INSURANCEDEKHO RESULT:")
    print(result_text)
    print("=" * 80)
    return result_text

async def main():
    """Run all insurance providers in parallel and save results to Excel"""
    print("\n" + "=" * 80)
    print("STARTING ALL 3 INSURANCE PROVIDERS IN PARALLEL")
    print("=" * 80)
    
    # Run all three providers in parallel using asyncio.gather
    results = await asyncio.gather(
        run_icici_lombard(),
        run_policybazaar(),
        run_insurancedekho(),
        return_exceptions=True
    )
    
    # Create Excel data with individual plan rows
    excel_data = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    providers = ["ICICI Lombard", "PolicyBazaar", "InsuranceDekho"]
    
    for i, provider in enumerate(providers):
        if isinstance(results[i], Exception):
            excel_data.append({
                "Provider": provider,
                "Vehicle": VEHICLE_REG_NO,
                "Plan Name": "Error",
                "Insurance Provider": provider,
                "Annual Premium (INR)": 0,
                "IDV": "",
                "Cashless Garages": "",
                "Key Benefits": str(results[i]),
                "Status": "Failed",
                "Timestamp": timestamp
            })
        else:
            # Try to parse JSON from result
            result_text = results[i]
            try:
                # Clean up the result text
                if "```json" in result_text:
                    result_text = result_text.split("```json")[1].split("```")[0].strip()
                elif "```" in result_text:
                    result_text = result_text.split("```")[1].split("```")[0].strip()
                
                # Remove escaped characters
                result_text = result_text.replace("\\n", "").replace('\\"', '"')
                if result_text.startswith('"""'):
                    result_text = result_text[3:-3].strip()
                
                # Try to parse as JSON
                data = json.loads(result_text)
                
                # Extract quotes/plans
                quotes = data.get("quotes", [])
                if quotes:
                    for quote in quotes:
                        # Join key benefits into readable text
                        benefits = quote.get("key_benefits", [])
                        benefits_text = "\n• " + "\n• ".join(benefits) if benefits else ""
                        
                        excel_data.append({
                            "Provider": provider,
                            "Vehicle": VEHICLE_REG_NO,
                            "Plan Name": quote.get("plan_name", "N/A"),
                            "Insurance Provider": quote.get("insurance_provider", provider),
                            "Annual Premium (INR)": quote.get("annual_premium_inr", 0),
                            "IDV": quote.get("idv", "N/A"),
                            "Cashless Garages": quote.get("cashless_garages", "N/A"),
                            "Key Benefits": benefits_text,
                            "Status": "Success",
                            "Timestamp": timestamp
                        })
                else:
                    # No quotes found, store raw result
                    excel_data.append({
                        "Provider": provider,
                        "Vehicle": VEHICLE_REG_NO,
                        "Plan Name": "No plans extracted",
                        "Insurance Provider": provider,
                        "Annual Premium (INR)": 0,
                        "IDV": "",
                        "Cashless Garages": "",
                        "Key Benefits": result_text[:500],
                        "Status": "Partial",
                        "Timestamp": timestamp
                    })
            except Exception as e:
                # JSON parsing failed, store raw result
                excel_data.append({
                    "Provider": provider,
                    "Vehicle": VEHICLE_REG_NO,
                    "Plan Name": "Parse Error",
                    "Insurance Provider": provider,
                    "Annual Premium (INR)": 0,
                    "IDV": "",
                    "Cashless Garages": "",
                    "Key Benefits": result_text[:500],
                    "Status": "Partial",
                    "Timestamp": timestamp
                })
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(excel_data)
    
    # Sort by Annual Premium for easy comparison
    df = df.sort_values(by="Annual Premium (INR)")
    
    filename = f"insurance_quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Insurance Quotes', index=False)
        
        # Get worksheet
        worksheet = writer.sheets['Insurance Quotes']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if max_length < cell_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Wrap text for Key Benefits column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=8, max_col=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    print("\n" + "=" * 80)
    print("RESULTS SAVED TO EXCEL")
    print("=" * 80)
    print(f"File: {filename}")
    print(f"\nSummary:")
    
    # Count successful extractions
    success_count = len([d for d in excel_data if d["Status"] == "Success"])
    print(f"  ✅ Successfully extracted {success_count} insurance plans")
    
    # Show premium range
    premiums = [d["Annual Premium (INR)"] for d in excel_data if d["Status"] == "Success" and d["Annual Premium (INR)"] > 0]
    if premiums:
        print(f"  💰 Premium range: ₹{min(premiums):,} - ₹{max(premiums):,}")
    
    # Provider summary
    for provider in providers:
        provider_plans = [d for d in excel_data if d["Provider"] == provider and d["Status"] == "Success"]
        status_icon = "✅" if provider_plans else "❌"
        print(f"  {status_icon} {provider}: {len(provider_plans)} plans")
    
    print("=" * 80)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except RuntimeError as e:
        if "Event loop is closed" in str(e):
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(main())
        else:
            raise e